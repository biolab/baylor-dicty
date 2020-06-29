import matplotlib.pyplot as plt
import glob
import pandas as pd
import numpy as np
from sklearn.preprocessing import minmax_scale
from skmultilearn.model_selection import iterative_train_test_split
import itertools
from sklearn.decomposition import PCA
from openpyxl import load_workbook
from scipy.stats import combine_pvalues

import arff

from networks.library_regulons import ClusterAnalyser, NeighbourCalculator, make_tsne, add_tsne, STRAIN_ORDER
from networks.functionsDENet import loadPickle, savePickle
from stages_DE.stages_library import *
from correlation_enrichment.library_correlation_enrichment import SimilarityCalculator

path = '/home/karin/Documents/timeTrajectories/data/deTime/de_time_impulse/'
lab = True
if lab:
    dataPath = '/home/karin/Documents/timeTrajectories/data/RPKUM/combined/'
    pathSelGenes = '/home/karin/Documents/timeTrajectories/data/regulons/selected_genes/'
    pathStages = '/home/karin/Documents/timeTrajectories/data/stages/'
    pathClassification = '/home/karin/Documents/timeTrajectories/data/stages/classification/'
    pathRegulons = '/home/karin/Documents/timeTrajectories/data/regulons/'
    pathReplicateImg = '/home/karin/Documents/timeTrajectories/data/replicate_image/'
    path_deOvR = '/home/karin/Documents/timeTrajectories/data/deTime/stage_vs_other/'
    path_de_neighbouring = '/home/karin/Documents/timeTrajectories/data/deTime/neighbouring/'

else:
    dataPath = '/home/karin/Documents/DDiscoideum/data/RPKUM/'

genes = pd.read_csv(dataPath + 'mergedGenes_RPKUM.tsv', sep='\t', index_col=0)
conditions = pd.read_csv(dataPath + 'conditions_mergedGenes.tsv', sep='\t', index_col=None)

# ******  How variable are  strains between replicates - plot variability as boxplots
# Stds of genes  in strains+timepoints (std of replicates for a gene) divided by mean
genes_conditions = ClusterAnalyser.merge_genes_conditions(genes=genes, conditions=conditions, matching='Measurment')
groups = genes_conditions.groupby(['Strain', 'Time'])
variation = {strain: [] for strain in conditions['Strain'].unique()}
for group in groups:
    name = group[0][0]
    data = group[1].drop(list(conditions.columns), axis=1)
    std = (data.std() / data.mean()).dropna()
    variation[name].extend(std)
plt.boxplot(list(variation.values()))
plt.gca().set_xticklabels(list(variation.keys()), rotation=90)

# *******************
# **** Dim reduction plot of samples

# ******** tSNE of measurements

averaged_data = pd.read_table(pathRegulons + 'genes_averaged_orange.tsv', index_col=0)

# Averaged or unaveraged data or averaged_AX4 data
# data=averaged_data[genes.index].T
# data=averaged_data.query('Strain =="AX4"')[genes.index]
data = genes[conditions.query('Strain =="AX4"')['Measurment']].copy().T

# Select genes - variable across stages in AX4 or not 0 - works better with non null genes
# Works better with all non null genes
selected_genes = data.T[(data != 0).any(axis=0)].index
# top_impulse = pd.read_table(pathReplicateImg + 'AX4_bestImpulse2000.tsv')
# selected_genes=top_impulse.Gene.values
data = data[selected_genes]

# Normalise data
names = data.index
gene_names = data.columns
# Works better with log2+m0s1 than minmax scaling
scaler = pp.StandardScaler()
data = pd.DataFrame(scaler.fit_transform(np.log2(data + 1)), index=names, columns=gene_names)
# tSNE
# tsne = make_tsne(data=data, perplexities_range=[50, 160], exaggerations=[1, 1], momentums=[0.6, 0.9], random_state=0)
tsne = make_tsne(data=data, perplexities_range=[8, 30], exaggerations=[1, 1], momentums=[0.6, 0.9], random_state=0)
# Data for plotting
# plot_data = pd.DataFrame(tsne, index=data.index, columns=['x', 'y'])

plot_data = pd.DataFrame()
# For embedding other strains onto AX4
mutants = list(conditions['Strain'].unique())
# mutants.remove('AX4')
for mutant in mutants:
    data = averaged_data.query('Strain =="' + mutant + '"')[selected_genes]
    print(mutant, data.shape)
    names = data.index
    gene_names = data.columns
    # Works better with log2+m0s1 than minmax scaling
    data = pd.DataFrame(scaler.transform(np.log2(data + 1)), index=names, columns=gene_names)
    # tSNE
    tsne2 = add_tsne(tsne1=tsne, data2=data)
    # Data for plotting
    plot_data = plot_data.append(pd.DataFrame(tsne2, index=data.index, columns=['x', 'y']))

# conditions_plot = conditions[['Replicate', 'Time', 'Group']]
# conditions_plot.index = conditions['Measurment']
# plot_data = pd.concat([plot_data, conditions_plot], axis=1)

# For unaveraged data
plot_data = pd.concat([plot_data,
                       pd.DataFrame(conditions.values, index=conditions['Measurment'].values,
                                    columns=conditions.columns.values)], axis=1, sort=True)
plot_by = 'Replicate'
# For averaged data
plot_data = pd.concat([plot_data, averaged_data[['Time', 'Strain']]], axis=1, sort=True)
plot_data['Group'] = [GROUPS[strain] for strain in plot_data['Strain']]
plot_by = 'Strain'

plot_data['size'] = minmax_scale(plot_data['Time'], (3, 30))

fig, ax = plt.subplots()
dim_reduction_plot(plot_data, plot_by=plot_by, fig_ax=(fig, ax), order_column='Time', colour_by_phenotype=False,
                   add_name=True)
ax.axis('off')
fig.suptitle("t-SNE of measurements. Size denotes time; replicate's progression is marked with a line.")

# ************************ PC 1 vs time of measurements on AX4 data
top_impulse = pd.read_table(pathReplicateImg + 'AX4_bestImpulse2000.tsv')

averaged_data = pd.read_table(pathRegulons + 'genes_averaged_orange.tsv', index_col=0)
X_avg_df_AX4 = averaged_data.query('Strain =="AX4"')[list(genes.index) + ['Time']]

# selected_genes=list(top_impulse.Gene.values)
selected_genes = X_avg_df_AX4.T[(X_avg_df_AX4 != 0).any(axis=0)].index
selected_genes = list(selected_genes.drop('Time'))

X_avg_df_AX4 = averaged_data.query('Strain =="AX4"')[selected_genes + ['Time']]

# Scale the data (sklearn's PCA does the centering itself)
# Scale data - log2 -y/n (also needs to be changed below in mutants) & m0s1 or minmax
# May work better with minmax no log?
scaler = pp.MinMaxScaler()
# X_avg_AX4 = scaler.fit_transform(np.log2(X_avg_df_AX4.drop('Time', axis=1)+1))
X_avg_AX4 = scaler.fit_transform(X_avg_df_AX4.drop('Time', axis=1))

pca = PCA(n_components=1, random_state=0)
pca_AX4 = pca.fit_transform(X_avg_AX4).ravel()

mutants = averaged_data['Strain'].unique().tolist()
mutants.remove('AX4')

plot_data = pd.DataFrame()
for strain in mutants:
    print(strain)
    X_avg_df = averaged_data.query('Strain =="' + strain + '"')[selected_genes + ['Time']]
    # X_avg = scaler.transform(np.log2(X_avg_df.drop('Time', axis=1)+1))
    X_avg = scaler.transform(X_avg_df.drop('Time', axis=1))
    pca_transformed = pca.transform(X_avg).ravel()
    plot_data = plot_data.append(
        pd.DataFrame({'x': X_avg_df['Time'], 'y': pca_transformed, 'size': [10] * X_avg_df.shape[0],
                      'Group': [GROUPS[strain]] * X_avg_df.shape[0], 'Strain': [strain] * X_avg_df.shape[0],
                      'width': [1.5] * X_avg_df.shape[0],
                      'alpha': [0.5] * X_avg_df.shape[0]}))

plot_data = plot_data.append(
    pd.DataFrame({'x': X_avg_df_AX4['Time'], 'y': pca_AX4, 'size': [30] * X_avg_df_AX4.shape[0],
                  'Group': ['WT'] * X_avg_df_AX4.shape[0], 'Strain': ['AX4'] * X_avg_df_AX4.shape[0],
                  'width': [4] * X_avg_df_AX4.shape[0],
                  'alpha': [1.0] * X_avg_df_AX4.shape[0]}))
plot_order = mutants + ['AX4']

fig, ax = plt.subplots()
dim_reduction_plot(plot_data, plot_by='Strain', fig_ax=(fig, ax), order_column='x', colour_by_phenotype=False,
                   add_name=True, legend_groups='upper left', fontsize=10, plot_order=plot_order)

ax.spines['right'].set_visible(False)
ax.spines['left'].set_visible(False)
ax.spines['top'].set_visible(False)
ax.set_xlabel('Time')
ax.set_ylabel('PC1')
ax.tick_params(axis='y', which='both', left=False, labelleft=False)
# ************************************************************************
# ***** Count in how many replicates per strain the gene is consistently 0
genes_rep = ClusterAnalyser.merge_genes_conditions(genes=genes, conditions=conditions[['Replicate', 'Measurment']],
                                                   matching='Measurment')
genes_rep = ClusterAnalyser.split_data(genes_rep, 'Replicate')
genes_zero_count = pd.DataFrame(np.zeros((genes.shape[0], conditions['Strain'].unique().shape[0])),
                                index=genes.index, columns=conditions['Strain'].unique())

for rep, data in genes_rep.items():
    data = data.drop(['Measurment', 'Replicate'], axis=1).T
    strain = conditions[conditions['Replicate'] == rep]['Strain'].values[0]
    data = (data == 0).all(axis=1)
    genes_zero_count[strain] = genes_zero_count[strain] + data
genes_zero_count.to_csv(dataPath + 'zero_replicates_count.tsv', sep='\t')
# ***************************
# **** Gene profile similarity to closest gene neighbours in strains
# Genes that are 0 in at least one replicate are removed from extraction of closest neighbours in that strain
NEIGHBOURS = 11
sims_dict = dict()

# Split data by strain, scaling and zero filtering is done in neighbours
merged = ClusterAnalyser.merge_genes_conditions(genes=genes, conditions=conditions[['Measurment', 'Strain']],
                                                matching='Measurment')
splitted = ClusterAnalyser.split_data(data=merged, split_by='Strain')
for strain, data in splitted.items():
    splitted[strain] = data.drop(["Strain", 'Measurment'], axis=1).T

# Calculate neighbours - sims_dict has similarity matrices from samples
for strain, data in splitted.items():
    print(strain)
    nonzero = genes_zero_count[genes_zero_count[strain] == 0].index
    data = data.loc[nonzero, :]
    neighbour_calculator = NeighbourCalculator(genes=data)
    neigh, sims_dict[strain] = neighbour_calculator.neighbours(n_neighbours=NEIGHBOURS, inverse=False,
                                                               scale='mean0std1',
                                                               log=True,
                                                               return_neigh_dist=True, remove_self=True)

savePickle(pathSelGenes + 'newGenes_noAll-removeSelf-removeZeroRep_simsDict_scalemean0std1_logTrue_kN' +
           str(NEIGHBOURS) + '_splitStrain.pkl', sims_dict)

# **************************
# ***** Find genes that are more correlated with close neighbours in some strains but not the others
sims_dict = loadPickle(
    pathSelGenes + 'newGenes_noAll-removeSelf-removeZeroRep_simsDict_scalemean0std1_logTrue_kN' + str(6) +
    '_splitStrain.pkl')
# Genes expressed in all strains:
# sel_genes = set(genes.index.values)
# for strain_data in sims_dict.values():
#    genes_strain = set(strain_data.index.values)
#    sel_genes = sel_genes & genes_strain

# Put 'avg similarity' of all 0 genes to lowest similarity of that strain (many genes should be all 0 but are not -
# erroneous mapping (not consistent shape between replicates). The low similarities may be results of unexpressed genes.
# Also do that for genes with at least one replicate all 0.
# Get means of gene similarities. If similarity was not calculated for a gene (because it had 0 expression), set it to
# min avg similarity. Do not set it to 0 as closest neighbours in individual strains are much above 0.
similarity_means = similarity_mean_df(sims_dict=sims_dict, index=genes.index, replace_na_sims=None,
                                      replace_na_mean='min')
# Get overall rank means
quantile_normalised = quantile_normalise(similarity_means=similarity_means)

quantile_normalised.to_csv(
    pathSelGenes + 'simsQuantileNormalised_newGenes_noAll-removeSelf-removeZeroRep_simsDict_scalemean0std1_logTrue_kN6_splitStrain.tsv',
    sep='\t')
# quantile_normalised=pd.read_table(pathSelGenes + 'simsQuantileNormalised_newGenes_noAll-removeSelf-removeZeroRep_simsDict_scalemean0std1_logTrue_kN6_splitStrain.tsv', index_col=0)

# # Fit sigmoids to find if similarities follow the shape (two plateaus) - not producing ok results
#
# fit_data = []
# could_not_fit = 0
# for gene in quantile_normalised.index:
#     x = []
#     y = []
#     for strain in quantile_normalised.columns:
#         group = GROUPS[strain]
#         if group in GROUP_X.keys():
#             x_pos = GROUP_X[group]
#             avg_similarity = quantile_normalised.loc[gene, strain]
#             x.append(x_pos)
#             y.append(avg_similarity)
#     try:
#         # Sometimes params can not be estimated
#         params = sigmoid_fit(x=x, y=y)[0]
#         cost = relative_cost(x=x, y=y, function=sigmoid, params=params)
#         fit_data.append({'Gene': gene, 'Cost': cost, 'L': params[0], 'x0': params[1], 'k': params[2], 'b': params[3]})
#     except:
#         could_not_fit = could_not_fit + 1
# fit_data = pd.DataFrame(fit_data)

# *** Mann–Whitney U or t test for comparing similarities distribution between strain groups
# Below: Tells which comparison on the strain developmental timeline to make
# First element group1, second group2, third comparison name
# TODO correct
group_splits = [
    ([1], [2, 3, 4, 6, 7], 1),
    ([1, 2], [3, 4, 6, 7], 2),
    ([1, 2, 3], [4, 6, 7], 3),
    ([1, 2, 3, 4], [6, 7], 4)
]

# results = compare_gene_scores(quantile_normalised=quantile_normalised, group_splits=group_splits,test='u',
#                             alternative='two-sided')
test = 't'
alternative = 'two-sided'
results = compare_gene_scores(quantile_normalised=quantile_normalised, group_splits=group_splits, test=test,
                              alternative=alternative)

results.to_csv(
    pathSelGenes + 'comparisonsAvgSims_' + test + '-' + alternative + '_newGenes_noAll-removeSelf-removeZeroRep_simsDict_scalemean0std1_logTrue_kN6_splitStrain.tsv',
    sep='\t', index=False)

# ***** Extract genes that are potential candidates for deregulation in some strains
data = pd.read_table(
    pathSelGenes + 'comparisonsAvgSimsSingle_AX4basedNeigh_u-less_newGenes_noAll-removeZeroRep_simsDict_scalemean0std1_logTrue_kN11_splitStrain.tsv',
    sep='\t')
# data['Max_mean'] = pd.concat([data['Mean1'], data['Mean2']], axis=1).max(axis=1)
filtered = data.query('FDR <=0.01 & Difference >=0.4 ')
filtered_genes = filtered['Gene'].unique()
genes_dict = dict(zip(filtered_genes, range(len(filtered_genes))))
comparisons = list(data['Comparison'].unique())
comparisons.sort()
comparisons_dict = dict(zip(comparisons, range(len(comparisons))))
comparison_df = np.zeros((filtered_genes.shape[0], len(comparisons)))
for row in filtered.iterrows():
    row = row[1]
    comparison_df[genes_dict[row['Gene']], comparisons_dict[row['Comparison']]] = 1
comparison_df = pd.DataFrame(comparison_df, index=filtered_genes, columns=['C' + str(c) for c in comparisons])
comparison_df.to_csv(
    pathSelGenes + 'summary_comparisonsAvgSimsSingle_AX4basedNeigh_u-less_newGenes_noAll-removeZeroRep_simsDict_scalemean0std1_logTrue_kN11_splitStrain.tsv',
    sep='\t')
# ***************
# # ** Compare close neighbours in AX4 with close neighbours in  other strains

# ***** Similarity to closest neighbours of AX4 across strains
# In each strain calculate similarity to genes that were identified as closest neighbours in AX4
# E. g. Do neighbourhoods change?
NEIGHBOURS = 11
sims_dict_WT = dict()
# Closest neighbours in AX4
strain = 'AX4'
nonzero = genes_zero_count[genes_zero_count[strain] == 0].index
data = splitted[strain]
data = data.loc[nonzero, :]
neighbour_calculator = NeighbourCalculator(genes=data)
neigh_WT, sims_dict_WT[strain] = neighbour_calculator.neighbours(n_neighbours=NEIGHBOURS, inverse=False,
                                                                 scale='mean0std1',
                                                                 log=True,
                                                                 return_neigh_dist=True, remove_self=True)
# Similarity in other strains
# For individual strains do not include genes that are all 0 in at least one replicate for calculation of any neighbours
for strain, data in splitted.items():
    if strain != 'AX4':
        print(strain)
        nonzero = set(genes_zero_count[genes_zero_count[strain] == 0].index)
        genes_WT = set(neigh_WT.index)
        nonzero = list(nonzero & genes_WT)
        data = data.loc[nonzero, :]
        data = pd.DataFrame(
            NeighbourCalculator.get_index_query(genes=data, inverse=False, scale='mean0std1', log=True)[0],
            index=data.index, columns=data.columns)
        n_genes = len(nonzero)
        similarities = np.empty((n_genes, NEIGHBOURS - 1))
        similarities[:] = np.nan
        for idx_gene in range(n_genes):
            gene = nonzero[idx_gene]
            neighbours_WT = neigh_WT.loc[gene, :].values
            for idx_neigh in range(NEIGHBOURS - 1):
                neigh = neighbours_WT[idx_neigh]
                if neigh in data.index:
                    similarities[idx_gene][idx_neigh] = SimilarityCalculator.calc_cosine(data.loc[gene, :],
                                                                                         data.loc[neigh, :])
        similarities = pd.DataFrame(similarities, index=nonzero, columns=range(NEIGHBOURS - 1))
        sims_dict_WT[strain] = similarities

savePickle(
    pathSelGenes + 'AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN' + str(NEIGHBOURS) +
    '_splitStrain.pkl', (neigh_WT, sims_dict_WT))

# ***** OPTION2 Similarity to closest neighbours of AX4 across strains
# Calculate neighbours on whole AX4 profile and then for similarities perform calculations on same n of samples of
# each strain, resampling multipole times to include more samples
# In each strain calculate similarity to genes that were identified as closest neighbours in AX4
# E. g. Do neighbourhoods change?
NEIGHBOURS = 11
# Closest neighbours in AX4
strain = 'AX4'
nonzero = genes_zero_count[genes_zero_count[strain] == 0].index
data = splitted[strain]
data = data.loc[nonzero, :]
neighbour_calculator = NeighbourCalculator(genes=data)
neigh_WT, sims_AX4 = neighbour_calculator.neighbours(n_neighbours=NEIGHBOURS, inverse=False,
                                                     scale='mean0std1',
                                                     log=True,
                                                     return_neigh_dist=True, remove_self=True)
# Similarity in other strains
# For individual strains do not include genes that are all 0 in at least one replicate for calculation of any neighbours
n_samples = 10
n_resample = 10
sims_dict_WT = dict()
random.seed(0)
genes_WT = set(neigh_WT.index)
for strain, data in splitted.items():
    print(strain)
    nonzero = set(genes_zero_count[genes_zero_count[strain] == 0].index)
    nonzero = list(nonzero & genes_WT)
    print('nonzero', len(nonzero))
    data = data.loc[nonzero, :]
    all_samples = list(data.columns)
    # Do for subsets of 10 samples
    similarities_resamples = []
    for i in range(n_resample):
        samples = random.sample(all_samples, n_samples)
        # Keep only data that is not all zero in samples subset
        data_sub = data.loc[:, samples]
        data_sub = data_sub[(data_sub != 0).any(axis=1)]
        print('sub', data_sub.shape[0], samples)
        data_sub = pd.DataFrame(
            NeighbourCalculator.get_index_query(genes=data_sub, inverse=False, scale='mean0std1', log=True)[0],
            index=data_sub.index, columns=data_sub.columns)
        n_genes = data_sub.shape[0]
        similarities = np.empty((n_genes, NEIGHBOURS - 1))
        similarities[:] = np.nan
        for idx_gene in range(n_genes):
            gene = data_sub.index[idx_gene]
            neighbours_WT = neigh_WT.loc[gene, :].values
            for idx_neigh in range(NEIGHBOURS - 1):
                neigh = neighbours_WT[idx_neigh]
                if neigh in data_sub.index:
                    similarities[idx_gene][idx_neigh] = SimilarityCalculator.calc_cosine(data_sub.loc[gene, :],
                                                                                         data_sub.loc[neigh, :])
        similarities = pd.DataFrame(similarities, index=data_sub.index, columns=range(NEIGHBOURS - 1))
        # Add genes for which similarities were not calculated
        similarities = similarities.reindex(genes_WT)
        # Replace non-calculated similarities (e.g. 0 expression) with 0
        similarities = similarities.replace(np.nan, 0)
        similarities_resamples.append(similarities.values)

    # Median of resampled similarities for each gene-neighbour pair is the saved similarity for the pair
    # Use median so that if profile was all 0 by chance and was thus given similarity 0 this will not affect it too much
    similarities_resamples = np.median(np.array(similarities_resamples), axis=0)
    similarities_resamples = pd.DataFrame(similarities_resamples, index=genes_WT, columns=range(NEIGHBOURS - 1))
    sims_dict_WT[strain] = similarities_resamples

savePickle(
    pathSelGenes + 'AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN' + str(NEIGHBOURS) +
    '_splitStrain_samples' + str(n_samples) + 'resample' + str(n_resample) + '.pkl', (neigh_WT, sims_dict_WT))

# ************ Find genes  for which neighbours in AX4 do not represent close neighbours (similar as above)
sims_dict_WT = loadPickle(
    pathSelGenes + 'AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN11_splitStrain.pkl')[1]

# Some neighbours and genes are nan as they had a replicate with all 0 expression.
# Replace np.nans in neighbours  (before mean calculation) or in gene means (query had all 0 replicate) with 0.
# Compared to above where neighbours from individual strains are used here the similarities to AX4 neighbours do drop
# to 0 or below so replacing nan with 0  will not lead to overly skewed (previously nan) values.
similarity_means_WT = similarity_mean_df(sims_dict=sims_dict_WT, index=genes.index, replace_na_sims=0,
                                         replace_na_mean='zero')
quantile_normalised_WT = quantile_normalise(similarity_means=similarity_means_WT)

# quantile_normalised_WT.to_csv(
#    pathSelGenes + 'simsQuantileNormalised_AX4basedNeigh_newGenes_noAll-removeZeroRep_simsDict_scalemean0std1_logTrue_kN11_splitStrain.tsv',
#    sep='\t')
# Pre select genes that are above a relative threshold in WT and PD strains (that develop)
genes_filtered = set(similarity_means_WT.index)
for strain in GROUP_DF[GROUP_DF['Group'].isin(['prec', 'WT'])]['Strain']:
    threshold = np.quantile(similarity_means_WT[strain], 0.3)
    genes_filtered = genes_filtered & set(similarity_means_WT[similarity_means_WT[strain] >= threshold].index)

test = 'u'
alternative = 'less'

group_splits = [
    ([1], [2, 3, 4, 5, 7, 8], 1),
    ([1, 2], [3, 4, 5, 7, 8], 2),
    ([1, 2, 3], [4, 5, 7, 8], 3),
    ([1, 2, 3, 4], [5, 7, 8], 4),
    ([1, 2, 3, 4, 5], [7, 8], 5)
]
select_single_comparsion = [[1, 2, 3, 4, 5, 7, 8], [1], [7, 8]]
group_df = GROUP_DF.copy()
# Dis together
group_df = GROUP_DF.copy()
group_df.loc[group_df['X'] == 2, 'X'] = 3
group_df.loc[group_df['X'] == 3, 'Group'] = 'dis'
group_splits = [
    ([1], [3, 4, 5, 7, 8], 1),
    ([1, 3], [4, 5, 7, 8], 3),
    ([1, 3, 4], [5, 7, 8], 4),
    ([1, 3, 4, 5], [7, 8], 5)
]
select_single_comparsion = [[1, 3, 4, 5, 7, 8], [1], [7, 8]]

# Test for all possible separation points
# results_WT = compare_gene_scores(quantile_normalised=quantile_normalised_WT.loc[genes_filtered, :],
#                                  group_splits=group_splits, test=test,
#                                  alternative=alternative)
# results_WT.to_csv(
#     pathSelGenes + 'comparisonsAvgSims_AX4basedNeigh_' + test + '-' + alternative + '_newGenes_noAll-removeZeroRep_simsDict_scalemean0std1_logTrue_kN11_splitStrain.tsv',
#     sep='\t', index=False)

# OR Test for single separation point
results_WT = compare_gene_scores(quantile_normalised=quantile_normalised_WT.loc[genes_filtered, :],
                                 group_splits=group_splits, test=test,
                                 alternative=alternative, select_single_comparsion=select_single_comparsion,
                                 comparison_selection='std', group_df=group_df)
results_WT.to_csv(
    pathSelGenes + 'comparisonsAvgSimsSingle2STDAny0.2_AX4basedNeigh_' + test + '-' + alternative + '_newGenes_noAll-removeZeroRep_simsDict_scalemean0std1_logTrue_kN11_splitStrain.tsv',
    sep='\t', index=False)

# ************ Find genes  for which neighbours in AX4 do not represent close neighbours in other strain groups, but
# do in WT
sims_dict = loadPickle(
    pathSelGenes + 'AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN11_splitStrain.pkl')[1]

# Make DF with all genes and neighbours in rows and strains in columns.
# Reindex individual strain data to have same gene index to start with.
# replace np.nan with 0
genes_AX4 = sims_dict['AX4'].index
merged_sims = pd.DataFrame()
for strain, data in sims_dict.items():
    data = data.reindex(genes_AX4)
    data = data.replace(np.nan, 0)
    # List all neighbours of all genes in a column, so that neighbours of a gene follow each other
    merged_sims[strain] = data.values.ravel()
merged_sims.index = np.array([[gene] * sims_dict['AX4'].shape[1] for gene in genes_AX4]).ravel()

# Quantile normalise across all neighbours and genes
quantile_normalised = quantile_normalise(similarity_means=merged_sims)
quantile_normalised.to_csv(
    pathSelGenes + 'simsQuantileNormalised_AX4basedNeigh_noAll-removeZeroRep_scalemean0std1_logTrue_kN11_splitStrain_nonAvg.tsv',
    sep='\t')

# Pre select genes that have WT neighbours median above regulons threshold
# Thresholds
threshold_dict_strain = pd.read_table(
    '/home/karin/Documents/timeTrajectories/data/regulons/selected_genes/thresholds/strainThresholds_k2_m0s1log_best0.7.tsv',
    sep='\t')
threshold_dict_strain = {data['Strain']: np.round(data['Threshold'], 3) for row, data in
                         threshold_dict_strain.iterrows()}

genes_filtered = set(genes_AX4)
for strain in GROUP_DF[GROUP_DF['Group'] == 'WT']['Strain']:
    threshold = threshold_dict_strain[strain]
    medians = sims_dict[strain].median(axis=1)
    genes_filtered = genes_filtered & set(medians[medians >= threshold].index)
    print(len(genes_filtered))

# Find genes without close neighbours in AX4
strains_WT = GROUP_DF[GROUP_DF['Group'] == 'WT']['Strain']
for group in GROUP_DF[GROUP_DF['Group'] != 'WT']['Group'].unique():
    print(group)
    strains = GROUP_DF[GROUP_DF['Group'] == group]['Strain']
    group_results = []
    for gene in genes_filtered:
        values_WT = quantile_normalised.loc[gene, strains_WT].values.ravel()
        values_mutant = quantile_normalised.loc[gene, strains].values.ravel()

        result = mannwhitneyu(values_mutant, values_WT, alternative='less')
        p = result[1]
        statistic = result[0]

        m_WT = values_WT.mean()
        m_mutant = values_mutant.mean()
        me_WT = np.median(values_WT)
        me_mutant = np.median(values_mutant)

        group_results.append({'Gene': gene, 'Statistic': statistic, 'p': p,
                              'Mean WT': m_WT, 'Mean mutant': m_mutant, 'Difference mean': m_WT - m_mutant,
                              'Median WT': me_WT, 'Median mutant': me_mutant, 'Difference median': me_WT - me_mutant})
    group_results = pd.DataFrame(group_results)
    # Adjust pvals to FDR
    group_results['FDR'] = multipletests(pvals=group_results['p'], method='fdr_bh', is_sorted=False)[1]
    group_results.to_csv(
        pathSelGenes + 'comparisonsSims_' + group + '_AX4basedNeigh_u-less_removeZeroRep_scalemean0std1_logTrue_kN11.tsv',
        sep='\t', index=False)

# ************ Find genes  for which neighbours in AX4 do not represent close neighbours in other strain groups, but
# do in WT; using the resampled data that does not need quantile normalisation - no difference in profile length
sims_dict = loadPickle(
    pathSelGenes + 'AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN11_splitStrain_samples10resample10.pkl')[
    1]

# Make DF with all genes and neighbours in rows and strains in columns.
# Reindexing and np replacement is not needed as this is done before
genes_AX4 = sims_dict['AX4'].index
merged_sims = pd.DataFrame()
for strain, data in sims_dict.items():
    if (data.index != genes_AX4).any() or data.isna().any().any():
        raise ValueError('Data index does not match AX4 index or data contains na values')
    # List all neighbours of all genes in a column, so that neighbours of a gene follow each other
    merged_sims[strain] = data.values.ravel()
merged_sims.index = np.array([[gene] * sims_dict['AX4'].shape[1] for gene in genes_AX4]).ravel()

merged_sims['Gene'] = merged_sims.index
medians = merged_sims.groupby('Gene').median()
merged_sims = merged_sims.drop('Gene', axis=1)
medians.to_csv(pathSelGenes +
               'simsMedian_AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN11_splitStrain_samples10resample10.tsv')

# Pre select genes that have WT neighbours above AX4 based threshold

genes_filtered = set(genes_AX4)
# Here quantile 0.2 and not 0.3 (regulons) is used - this gives similar graph threshold as in regulons
# This may be because regulons were computed on genes not null in AX4 (N=12316) and these similarities only
# on genes non-null in any AX4 replicate (N=11501), thus including less genes
# In both AX4 and MybBGFP N=6171 (in AX4: 9201, MybBGFP: 6414) genes were selected at threshold 0.7743873008057314
threshold = medians['AX4'].quantile(0.2)
for strain in GROUP_DF[GROUP_DF['Group'] == 'WT']['Strain']:
    medians_strain = medians[strain]
    selected_genes_strain=set(medians_strain[medians_strain >= threshold].index)
    genes_filtered = genes_filtered & selected_genes_strain
    print('selected in',strain,len(selected_genes_strain),'remaining',len(genes_filtered))


# Find genes without close neighbours in AX4
strains_WT = GROUP_DF[GROUP_DF['Group'] == 'WT']['Strain']
for group in GROUP_DF[GROUP_DF['Group'] != 'WT']['Group'].unique():
    print(group)
    strains = GROUP_DF[GROUP_DF['Group'] == group]['Strain']
    group_results = []
    for gene in genes_filtered:
        values_WT = merged_sims.loc[gene, strains_WT].values.ravel()
        values_mutant = merged_sims.loc[gene, strains].values.ravel()

        result = mannwhitneyu(values_mutant, values_WT, alternative='less')
        p = result[1]
        statistic = result[0]

        m_WT = values_WT.mean()
        m_mutant = values_mutant.mean()
        me_WT = np.median(values_WT)
        me_mutant = np.median(values_mutant)

        group_results.append({'Gene': gene, 'Statistic': statistic, 'p': p,
                              'Mean WT': m_WT, 'Mean mutant': m_mutant, 'Difference mean': m_WT - m_mutant,
                              'Median WT': me_WT, 'Median mutant': me_mutant, 'Difference median': me_WT - me_mutant})
    group_results = pd.DataFrame(group_results)
    # Adjust pvals to FDR
    group_results['FDR'] = multipletests(pvals=group_results['p'], method='fdr_bh', is_sorted=False)[1]
    group_results.to_csv(
        pathSelGenes + 'comparisonsSims_' + group + '_AX4basedNeigh_u-less_removeZeroRep_scalemean0std1_logTrue_kN11_samples10resample10.tsv',
        sep='\t', index=False)

# ****** For each strain/gene compare similarities to neighbours from AX4 and closest neighbours in the strain
# Similarities to neighbours from AX4 - do not quantile normalise !!!!!
sims_dict_WT = loadPickle(
    pathSelGenes + 'AX4basedNeigh_newGenes-removeZeroRep_neighSimsDict_scalemean0std1_logTrue_kN11_splitStrains.pkl')[1]
similarity_means_WT = similarity_mean_df(sims_dict=sims_dict_WT, index=genes.index, replace_na_sims=0,
                                         replace_na_mean=None)

# Similarities to neighbours from individual strains - do not quantile normalise as changes scale compared to above
sims_dict = loadPickle(
    pathSelGenes + 'newGenes_noAll-removeSelf-removeZeroRep_simsDict_scalemean0std1_logTrue_kN11_splitStrain.pkl')

similarity_means = similarity_mean_df(sims_dict=sims_dict, index=genes.index, replace_na_sims=None,
                                      replace_na_mean=None)
threshold_dict_strain = dict()
for strain in similarity_means.columns:
    threshold_dict_strain[strain] = np.quantile(similarity_means[strain][similarity_means[strain].notna()], 0.3)

genes_AX4 = set(similarity_means[similarity_means['AX4'] >= threshold_dict_strain['AX4']].index)
diffs_df = pd.DataFrame()
for strain in conditions['Strain'].unique():
    # Has genes non-rep-zero in strain
    # Has genes non-rep-zero in strain and AX4 (less genes than sims_strain)
    strain_genes = set(sims_dict_WT[strain].index)
    strain_genes = list(strain_genes & genes_AX4)
    x = GROUP_DF[GROUP_DF['Strain'] == strain]['X'].values[0]
    n_genes = len(strain_genes)
    strain_means = similarity_means.loc[strain_genes, strain].values
    high = strain_means >= threshold_dict_strain[strain]
    diffs_df = pd.concat(
        [diffs_df, pd.DataFrame({'Gene': strain_genes, 'Strain': [strain] * n_genes, 'X': [x] * n_genes,
                                 'Sim_AX4': similarity_means_WT.loc[strain_genes, strain].values,
                                 'Sim_strain': strain_means, 'High': high})])

diffs_df.to_csv(pathSelGenes + 'diffs_AX4Strain.tsv', sep='\t', index=False)

# *******************************
# ***** Add phenotype info to conditions
conditions = conditions.drop(PHENOTYPES, axis=1)
conditions = pd.concat([conditions, pd.DataFrame(np.zeros((conditions.shape[0], len(PHENOTYPES))), columns=PHENOTYPES)],
                       axis=1)
# How to fill cells with no image (0 or -1 (for plotting/avg summary)):
no_image_fill = -1

no_seq = 0
no_image = 0
annotated = 0
file = '/home/karin/Documents/timeTrajectories/data/from_huston/phenotypes/Phenotype_milestone_strains.xlsx'
wb = load_workbook(file, read_only=True)

# files = [f for f in glob.glob('/home/karin/Documents/timeTrajectories/data/from_huston/phenotypes/' + "*.tab")]
# for f in files:
# phenotypes = pd.read_table(f, index_col=0)
for strain in conditions['Strain'].unique():
    if strain in wb.sheetnames:
        phenotypes = pd.read_excel(file, sheet_name=strain, index_col=0)
        phenotypes = phenotypes.replace('no image', np.nan)
        for time in phenotypes.index:
            for replicate in phenotypes.columns:
                val = phenotypes.loc[time, replicate]

                # Find conditions row of replicate+time
                conditions_row = conditions[(conditions['Replicate'] == replicate) & (conditions['Time'] == time)]
                if conditions_row.shape[0] > 0:
                    idx_name_conditions = \
                        conditions[(conditions['Replicate'] == replicate) & (conditions['Time'] == time)].index[0]

                    # Check if there is a phenotype (str) or no phenotype annotation
                    if type(val) == str:
                        val = val.replace('rippling', 'stream')
                        val = val.replace('strem', 'stream')
                        val = val.replace('no agg', 'no_agg')
                        val = val.replace('noAgg', 'no_agg')
                        val = val.replace('small', '')
                        val = val.replace(')', '')
                        val = val.replace('(', '')
                        val = val.replace(' ', '')
                        val = val.replace('elongating', '')
                        val = val.replace('shrinking', '')
                        val = val.replace('Spore', '_spore')
                        val = val.replace('streaming', 'stream')
                        val = val.split('/')
                        val = list(filter(lambda a: a != '', val))
                        annotated += 1
                        for pheno in val:
                            if pheno not in PHENOTYPES:
                                # print(f, pheno, time, replicate)
                                print(strain, pheno, time, replicate)
                            else:
                                conditions.at[idx_name_conditions, pheno] = 1
                    # Add -1 (or 0) to all phenotypes if there is no image for that time
                    else:
                        no_image += 1
                        conditions.loc[idx_name_conditions, PHENOTYPES] = [no_image_fill] * len(PHENOTYPES)
                        # print(val, replicate, time)

                else:
                    if type(val) == str:
                        print('No sample for', replicate, time)
                        no_seq += 1
    else:
        print('Strain not in file:', strain)
        for idx_name_conditions in conditions.query('Strain =="' + strain + '"').index:
            no_image += 1
            conditions.loc[idx_name_conditions, PHENOTYPES] = [no_image_fill] * len(PHENOTYPES)

if no_image_fill == -1:
    conditions.to_csv(dataPath + 'conditions_noImage_mergedGenes.tsv', sep='\t', index=False)
elif no_image_fill == 0:
    # Make all 0 times no_agg if not already filled
    phenotypes_notnoag = PHENOTYPES.copy()
    phenotypes_notnoag.remove('no_agg')
    for idx, sample in conditions.iterrows():
        if sample['Time'] < 1:
            if not (sample[PHENOTYPES] > 0).any():
                conditions.at[idx, 'no_agg'] = 1
                conditions.loc[idx, phenotypes_notnoag] = [0] * len(phenotypes_notnoag)
                annotated += 1
                no_image -= 1
            # else:
            #    print(sample)

    conditions.to_csv(dataPath + 'conditions_mergedGenes.tsv', sep='\t', index=False)

# *******************
# **** Add main stage to conditions
conditions['main_stage'] = ''

no_seq = 0
no_main = 0
no_image_strain = 0
annotated = 0
file = '/home/karin/Documents/timeTrajectories/data/from_huston/phenotypes/main_stage.xlsx'
wb = load_workbook(file, read_only=True)

for strain in conditions['Strain'].unique():
    if strain in wb.sheetnames:
        phenotypes = pd.read_excel(file, sheet_name=strain, index_col=0)
        phenotypes = phenotypes.replace('no image', np.nan)
        phenotypes = phenotypes.replace('no_image', np.nan)
        phenotypes = phenotypes.replace('unsynchronized', np.nan)
        for time in phenotypes.index:
            for replicate in phenotypes.columns:
                val = phenotypes.loc[time, replicate]

                # Find conditions row of replicate+time
                conditions_row = conditions[(conditions['Replicate'] == replicate) & (conditions['Time'] == time)]
                if conditions_row.shape[0] > 0:
                    idx_name_conditions = \
                        conditions[(conditions['Replicate'] == replicate) & (conditions['Time'] == time)].index[0]

                    # Check if there is a phenotype (str) or no phenotype annotation
                    if type(val) == str:
                        val = val.replace('rippling', 'stream')
                        val = val.replace('strem', 'stream')
                        val = val.replace('no agg', 'no_agg')
                        val = val.replace('noAgg', 'no_agg')
                        val = val.replace('small', '')
                        val = val.replace(')', '')
                        val = val.replace('(', '')
                        val = val.replace(' ', '')
                        val = val.replace('elongating', '')
                        val = val.replace('shrinking', '')
                        val = val.replace('Spore', '_spore')
                        val = val.replace('streaming', 'stream')
                        annotated += 1
                        if val not in PHENOTYPES:
                            print(strain, val, time, replicate)
                        else:
                            conditions.at[idx_name_conditions, 'main_stage'] = val
                    else:
                        no_main += 1
                        conditions.loc[idx_name_conditions, 'main_stage'] = None
                        # print(val, replicate, time)

                else:
                    if type(val) == str:
                        print('No sample for', replicate, time)
                        no_seq += 1
    else:
        # print('Strain not in file:', strain)
        for idx_name_conditions in conditions.query('Strain =="' + strain + '"').index:
            no_image_strain += 1
            # conditions.loc[idx_name_conditions, 'main_stage'] = ''

# Make all 0 times no_agg if not already filled. Do this only for strains that have also other annotations.
for idx, sample in conditions.iterrows():
    if sample['Time'] == 0:
        if (sample['main_stage'] == None):
            conditions.at[idx, 'main_stage'] = 'no_agg'
            annotated += 1
            no_main -= 1
        # else:
        #    print(sample['Short'])

conditions['main_stage'] = conditions['main_stage'].fillna(np.nan).replace('', np.nan)

conditions.to_csv(dataPath + 'conditions_mergedGenes.tsv', sep='\t', index=False)

# ***********************************************
# ****** Averaged stages - for a timepoint add all stages present in each replicate
phenotypes = ['no_agg', 'stream', 'lag', 'tag', 'tip', 'slug', 'mhat', 'cul', 'FB', 'yem']


def set_infered(averaged: dict, infered_stages: list):
    for col in phenotypes:
        fill = 'no'
        if col in infered_stages:
            fill = 'no image'
        averaged[col] = fill


def avgsample_name(group):
    return group[0] + '_' + str(group[1])


conditions_noimg = pd.read_csv(dataPath + 'conditions_noImage_mergedGenes.tsv', sep='\t',
                               index_col=None).sort_values(['Strain', 'Time'])
averaged_stages = pd.DataFrame(columns=phenotypes)
grouped = conditions_noimg.groupby(['Strain', 'Time'])
for group, data in grouped:
    name = avgsample_name(group)
    averaged = {}
    # max_rep=conditions.query('Strain == "'+group[0]+'"')['Replicate'].unique().shape[0]
    for col in phenotypes:
        pheno_data = data[col]
        if (pheno_data == 1).any():
            averaged[col] = 'yes'
        # Decide if combination of 0 an -1 is unknown or known - there could be this phenotype in the sample without image
        # Use this instead of the below to put unknown only if all are unknown (so not if some are 'no')
        elif (pheno_data == -1).all():
            averaged[col] = 'no image'
        else:
            averaged[col] = 'no'

        # If any unknown or less than all replicates put unknown - NOT as we do not have expression for them either
        # elif (pheno_data == 0).all() and pheno_data.shape[0]==max_rep:
        # elif (pheno_data == 0).all():
        #     averaged[col] = 'no'
        # else:
        #     averaged[col] = 'unknown'

    # Modify data with inferred stages if there is no image at that time point (except for strains with no data at all)
    time = group[1]
    strain = group[0]
    # Not strain=='ac3PkaCoe' and time==0 as this does not remove this strain from other comparisons
    if strain == 'ac3PkaCoe':
        if time == 0:
            set_infered(averaged=averaged, infered_stages=['no_agg'])
    elif strain == 'gtaC':
        set_infered(averaged=averaged, infered_stages=['no_agg'])
    else:
        if (np.array(list(averaged.values())) == 'no image').all():
            if time == 0:
                set_infered(averaged=averaged, infered_stages=['no_agg'])
            else:
                strain_times = conditions_noimg.query('Strain =="' + strain + '"')['Time'].unique()
                strain_times.sort()
                previous_time = int(strain_times[np.argwhere(strain_times == time) - 1])
                previous_data = averaged_stages.loc[avgsample_name((strain, previous_time)), :]
                # Previous time not a definite no - set also this time to this
                previous_stages = list(previous_data[previous_data != 'no'].index)
                set_infered(averaged=averaged, infered_stages=previous_stages)

    averaged_stages = averaged_stages.append(pd.DataFrame(averaged, index=[name]), sort=True)
averaged_stages.index.name = 'Name'
averaged_stages = averaged_stages.reindex(phenotypes, axis=1)
averaged_stages.to_csv(pathStages + 'averageStages.tsv', sep='\t')

# *********************
# ******** Find genes overexpressed in a stage (data from R deSeq2 1 vs 1 stage)
files = [f for f in glob.glob('/home/karin/Documents/timeTrajectories/data/deTime/stage_vs_stage/' + "*.tsv")]
# Remove disappear as this will be analysed separatley
files = [f for f in files if 'disappear' not in f]
stage_genes = pd.DataFrame()
for stage in PHENOTYPES:
    files_stage = [f for f in files if stage in f.split('/')[-1]]
    selected_genes = set(genes.index)
    print('*****', stage, len(files_stage))
    if len(files_stage) > 0:
        for file in files_stage:
            # print(file)
            data = pd.read_table(file, index_col=0)
            file_fields = file.split('ref')
            if stage in file_fields[0]:
                direction = '>= '
            elif stage in file_fields[1]:
                direction = '<= -'
            # Filter based on fold change to get overexpressed
            markers = set(data.query('log2FoldChange ' + direction + '1').index)
            selected_genes = selected_genes & markers
            # print(len(selected_genes))
        print(len(selected_genes))
        for gene in selected_genes:
            stage_genes.loc[gene, stage] = 1

stage_genes.to_csv('/home/karin/Documents/timeTrajectories/data/deTime/stage_vs_stage/markers.tab', sep='\t')

# ******************************************************
# ********** Genes that always peak in certain stage

# Find for each gene a peak time in each replicate. Leave nan if gene constantly 0.
merged = ClusterAnalyser.merge_genes_conditions(genes=genes, conditions=conditions[['Measurment', 'Replicate']],
                                                matching='Measurment')
splitted = ClusterAnalyser.split_data(data=merged, split_by='Replicate')
for strain, data in splitted.items():
    splitted[strain] = data.drop(["Replicate", 'Measurment'], axis=1).T

genes_dict = dict(zip(genes.index, range(genes.shape[0])))
replicates = list(splitted.keys())
reps_dict = dict(zip(replicates, range(len(replicates))))
peak_data = np.empty((genes.shape[0], len(replicates)))
peak_data[:] = np.nan
conditions.index = conditions['Measurment']
for rep, data in splitted.items():
    print(rep)
    data = data.copy()
    data.columns = conditions.loc[data.columns]['Time'].values
    for gene, expression in data.iterrows():
        if not (expression == 0).all():
            peak_data[genes_dict[gene]][reps_dict[rep]] = ClusterAnalyser.peak(expression)

peak_data = pd.DataFrame(peak_data, index=genes.index, columns=replicates)
peak_data.to_csv(pathStages + 'peaks.tsv', sep='\t')

# Determine for each replicate in which stage(s) had each gene a peak
peak_data = pd.read_table(pathStages + 'peaks.tsv', index_col=0)

genes_dict = dict(zip(genes.index, range(genes.shape[0])))
phenotypes_dict = dict(zip(PHENOTYPES, range(len(PHENOTYPES))))
peak_counts = dict()

for replicate in conditions['Replicate'].unique():
    peak_stage = np.zeros((len(genes_dict), len(phenotypes_dict)))
    print(replicate)
    stage_data = conditions[conditions['Replicate'] == replicate][PHENOTYPES]
    stage_data.index = conditions[conditions['Replicate'] == replicate]['Time']
    for gene, peak in peak_data[replicate].iteritems():
        if not np.isnan(peak):
            stages = stage_data.loc[peak, :]
            for stage, present in stages.iteritems():
                if present == 1:
                    peak_stage[genes_dict[gene]][phenotypes_dict[stage]] += 1
    peak_counts[replicate] = pd.DataFrame(peak_stage, index=genes.index, columns=PHENOTYPES)

# Merge for all replicates - Count in how many replicates  had a gene a peak in certain stage
combined_counts = pd.DataFrame(np.zeros((len(genes_dict), len(phenotypes_dict))), index=genes.index, columns=PHENOTYPES)
for data in peak_counts.values():
    combined_counts = combined_counts + data

# Convert counts of peak to proportions (divide by N of replicates that have that stage)
combined_proportion = pd.DataFrame()
for phenotype in combined_counts.columns:
    n = conditions[conditions[phenotype] == 1]['Replicate'].unique().shape[0]
    combined_proportion[phenotype] = combined_counts[phenotype] / n
combined_proportion.to_csv(pathStages + 'stage_peak_proportion.tsv', sep='\t')
# Problem as misses genes that have a peak in stage without image

# Only Wt data
combined_counts_WT = pd.DataFrame(np.zeros((len(genes_dict), len(phenotypes_dict))), index=genes.index,
                                  columns=PHENOTYPES)
for replicate in conditions[conditions['Group'] == 'WT']['Replicate'].unique():
    combined_counts_WT = combined_counts_WT + peak_counts[replicate]

combined_proportion_WT = pd.DataFrame()
for phenotype in combined_counts_WT.columns:
    n = conditions.query(phenotype + ' == 1 & Group =="WT"')['Replicate'].unique().shape[0]
    combined_proportion_WT[phenotype] = combined_counts_WT[phenotype] / n
combined_proportion_WT.to_csv(pathStages + 'stage_peak_proportion_WT.tsv', sep='\t')

# Compute random null distribution for N (or % in each stage) by shuffling counts within stages for each replicate
# (stage gets same number of genes that peak in it, but they are distributed randomly)
# Then do other steps (summing across replicates, proportion of replicates) to get the null distribution for each stage
peak_counts_permuted = dict()
for replicate, data in peak_counts.items():
    data_shuffled = pd.DataFrame(index=data.index, columns=data.columns)
    for col in data.columns:
        col_copy = data[col].copy().values
        np.random.shuffle(col_copy)
        data_shuffled[col] = col_copy
    peak_counts_permuted[replicate] = data_shuffled

# For WT and all strains the differences in distn are so big that it is not relevant (e.g. threshold could be at
# 0.5/0.6) (tested on tag)

# *******************************************
# ********** Find genes that best predict phenotype (classification based)

# ********* Prepare train test split

# Uses all data that has something annotated (either from images or as it was 0 time -noagg)
# Removes tag_spore - uses even samples that now have all 0 stage annotations if they had previously only tag_spore
Y = conditions[(conditions[PHENOTYPES] != 0).any(axis=1)]
X = genes[Y.Measurment].T.values
# PHENOTYPES here also ensures correct order
order = PHENOTYPES.copy()
order.remove('tag_spore')
Y = Y[order].values

# Save order of genes and phenotypes for classification
savePickle(pathClassification + 'feature_order.pkl', list(genes.index))
savePickle(pathClassification + 'target_order.pkl', order)

# Make train/test split for multilabel classification
X_train, Y_train, X_test, Y_test = iterative_train_test_split(X, Y, test_size=0.1)
# Split
for col in range(Y.shape[1]):
    print(Y_train[:, col].sum(), Y_test[:, col].sum())
print(Y_train.shape[0], Y_test.shape[0])
savePickle(pathClassification + 'train_test.pkl', (X_train, Y_train, X_test, Y_test))

# ********* Prepare data for Clus

# Save data in arff format for Clus
# Load data
X_train, Y_train, X_test, Y_test = loadPickle(pathClassification + 'train_test.pkl')
features_order = loadPickle(pathClassification + 'feature_order.pkl')
X_train = pd.DataFrame(X_train, columns=features_order)
target_order = loadPickle(pathClassification + 'target_order.pkl')
Y_train = pd.DataFrame(Y_train, columns=target_order)
Y_train = Y_train.astype('str')
data = pd.concat([Y_train, X_train], axis=1)

# Format data
arff_data = {}
arff_data['relation'] = 'stages'
attributes = []
for col in data.columns:
    type = None
    if col in target_order:
        type = list(data[col].unique())
    elif col in features_order:
        type = 'NUMERIC'
    else:
        raise ValueError('Unknown feature')
    attributes.append((col, type))
arff_data['attributes'] = attributes
arff_data['data'] = data.values.tolist()

# Save data
with open(pathClassification + 'train.arff', 'w') as f:
    arff.dump(arff_data, f)


# Prepare settings file for clus
def write_line(file, line):
    file.write(line + '\n')


minimalWeight_list = [3, 10]
iterations_list = [10, 50, 100]
selectRandomSubspaces_list = [1000, 4000, 10000]
params_grid = list(itertools.product(*[minimalWeight_list, iterations_list, selectRandomSubspaces_list]))
for params in params_grid:
    minimalWeight = params[0]
    iterations = params[1]
    selectRandomSubspaces = params[2]
    with open(pathClassification + 'clus/stages.s', 'w') as f:
        write_line(f, '[Data]')
        write_line(f, 'File = ' + pathClassification + 'train.arff')
        write_line(f, 'XVal = 5')
        write_line(f, '')
        write_line(f, '[Attributes]')
        write_line(f, 'Target = 1-' + str(len(target_order)))
        write_line(f, '')
        write_line(f, '[Model]')
        write_line(f, 'MinimalWeight = ' + str(minimalWeight))
        write_line(f, '')
        write_line(f, '[Ensemble]')
        write_line(f, 'Iterations = ' + str(iterations))
        write_line(f, 'EnsembleMethod = RForest')
        write_line(f, 'SelectRandomSubspaces = ' + str(selectRandomSubspaces))
        write_line(f, 'PrintAllModels = Yes')
        # write_line(f, 'OOBestimate = Yes')
        # write_line(f, 'FeatureRanking = Yes')
        write_line(f, '')
        write_line(f, '[Output]')
        write_line(f, 'ValidErrors = Yes')
        write_line(f, 'WriteErrorFile = Yes')
        write_line(f, 'WritePredictions = {Test,Train}')

# Run clus
# java -jar /home/karin/Documents/Clus/Clus.jar -xval -forest /home/karin/Documents/timeTrajectories/data/stages/classification/clus/stages.s  > stages_out.txt  2>stages_error.txt

# ********************************
# ******** Parse DE one vs rest results into a single file
# folder='nobatchrep'
# folder = 'WT_batchrep'
folder = 'strain_batchrep_lFC05_upper'
# Files for normal stages or strain-combined stages
# files = [f for f in glob.glob(path_deOvR + folder + '/' + "*.tsv")]
files = [f for f in glob.glob(path_deOvR + folder + '/' + "*_combined.tsv")]
combined = pd.DataFrame()
for stage in PHENOTYPES:
    # File for normal stages or strain-combined stages
    # file = path_deOvR + folder + '/DE_' + stage + '_ref_other_padj0.05_lFC1.tsv'
    file = path_deOvR + folder + '/' + stage + '_combined.tsv'
    if file in files:
        # print(stage)
        data = pd.read_table(file, index_col=0)
        data['Stage'] = stage
        combined = pd.concat([combined, data], sort=True)
# Only for the combined starins
col_order = [col for col in combined.columns if col not in ['combined_pval', 'pval_n', 'combined_padj', 'Stage']]
col_order = ['combined_pval', 'pval_n', 'combined_padj', 'Stage'] + col_order
combined = combined[col_order]

combined.to_csv(path_deOvR + folder + '_combined.tsv', sep='\t')

# *************************
# ***** Combine p values of DESeq2 from different strains
for stage in PHENOTYPES:
    files = [f for f in
             glob.glob(path_deOvR + 'strain_batchrep_lFC05_upper/*_DE_' + stage + '_ref_other_padj_lFC' + "*.tsv")]
    print(stage, len(files))
    # if len(files)>1:
    combined_stage = pd.DataFrame()
    for file in files:
        strain = file.split('/')[-1].split('_')[0]
        data = pd.read_table(file, index_col=0)
        data.columns = [strain + '_' + col for col in data.columns]
        combined_stage = pd.concat([combined_stage, data], axis=1, sort=True)
    # First add new pval cols before selecting them, make sure new ones are not included in analysisis
    combined_stage['combined_pval'] = np.nan
    combined_stage['pval_n'] = np.nan
    pval_cols = [col for col in combined_stage.columns if 'pval' in col and col not in ['combined_pval', 'pval_n']]
    for gene, data in combined_stage.iterrows():
        data = data[pval_cols]
        if (~np.isnan(data)).sum() > 1:
            pvals = data[~np.isnan(data)].values
            pval = combine_pvalues(pvalues=pvals, method='stouffer')[1]
            combined_stage.loc[gene, ['combined_pval', 'pval_n']] = [pval, len(pvals)]
        elif (~np.isnan(data)).sum() == 1:
            combined_stage.loc[gene, ['combined_pval', 'pval_n']] = [data[~np.isnan(data)], len(pvals)]
            combined_stage.loc[gene, 'pval_n'] = 1
    not_all = combined_stage.query('pval_n < ' + str(len(files))).copy()
    is_all = combined_stage.query('pval_n == ' + str(len(files))).copy()
    is_all['combined_padj'] = multipletests(is_all['combined_pval'], method='fdr_bh')[1]
    combined_stage = pd.concat([is_all, not_all], sort=False)
    combined_stage.to_csv(path_deOvR + 'strain_batchrep_lFC05_upper/' + stage + '_combined.tsv', sep='\t')

# ************************************
# *** Save non null genes
genes_nn = pd.DataFrame(genes[(genes != 0).any(axis=1)].index.values, columns=['Gene'])
genes_nn.to_csv(dataPath + 'nonNullGenes.tsv', sep='\t', index=False)

# *************
# *** Average data for main stages
# Select for now only WT as only this has all data and samples that have annotated main stage
# conditions_main = conditions.query('Group =="WT"').query('~main_stage.isna()', engine='python')
conditions_main = conditions.query('~main_stage.isna()', engine='python')
# Average expression
genes_group = genes[conditions_main.Measurment].copy().T
genes_group = genes_group.join(pd.DataFrame(conditions_main[['Strain', 'main_stage']].values,
                                            index=conditions_main['Measurment'], columns=['Strain', 'main_stage']))
averaged = genes_group.groupby(['Strain', 'main_stage']).mean().reset_index()
# Sort by strain and stage
averaged['main_stage'] = pd.Categorical(averaged['main_stage'], categories=PHENOTYPES, ordered=True)
averaged['Strain'] = pd.Categorical(averaged['Strain'], categories=STRAIN_ORDER, ordered=True)
averaged = averaged.sort_values(['Strain', 'main_stage'])

averaged.to_csv(pathStages + 'genes_averaged_orange_mainStage.tsv', sep='\t')

# Scale
percentile = 0.99
max_val = 0.1
genes_orange_avg_scaled = averaged.copy()
genes_orange_avg_scaled = genes_orange_avg_scaled.drop(['Strain', 'main_stage'], axis=1)
genes_avg_percentile = genes_orange_avg_scaled.quantile(q=percentile, axis=0)
genes_orange_avg_scaled = genes_orange_avg_scaled - genes_avg_percentile
genes_avg_percentile = genes_avg_percentile.replace(0, 1)
genes_orange_avg_scaled = genes_orange_avg_scaled / genes_avg_percentile
# Bound at max 0.1 (to remove outliers)
genes_orange_avg_scaled[genes_orange_avg_scaled > max_val] = max_val
genes_orange_avg_scaled[['Strain', 'main_stage']] = averaged[['Strain', 'main_stage']]
genes_orange_avg_scaled['Group'] = [GROUPS[strain] for strain in genes_orange_avg_scaled['Strain']]
genes_orange_avg_scaled.index = [strain + '_' + main
                                 for strain, main in genes_orange_avg_scaled[['Strain', 'main_stage']].values]
genes_orange_avg_scaled.to_csv(pathStages + 'genes_averaged_orange_mainStage_scale' + str(percentile)[2:] +
                               'percentileMax' + str(max_val) + '.tsv', sep='\t')

# *************
# *** Average data for main stages
# Select for now only WT as only this has all data and samples that have annotated main stage
# conditions_main = conditions.query('Group =="WT"').query('~main_stage.isna()', engine='python')
conditions_main = conditions.query('~main_stage.isna()', engine='python')
# Average expression
genes_group = genes[conditions_main.Measurment].copy().T
genes_group = genes_group.join(pd.DataFrame(conditions_main[['Strain', 'main_stage']].values,
                                            index=conditions_main['Measurment'], columns=['Strain', 'main_stage']))
averaged = genes_group.groupby(['Strain', 'main_stage']).mean().reset_index()
# Sort by strain and stage
averaged['main_stage'] = pd.Categorical(averaged['main_stage'], categories=PHENOTYPES, ordered=True)
averaged['Strain'] = pd.Categorical(averaged['Strain'], categories=STRAIN_ORDER, ordered=True)
averaged = averaged.sort_values(['Strain', 'main_stage'])

averaged.to_csv(pathStages + 'genes_averaged_orange_mainStage.tsv', sep='\t')

# Scale
percentile = 0.99
max_val = 0.1
genes_orange_avg_scaled = averaged.copy()
genes_orange_avg_scaled = genes_orange_avg_scaled.drop(['Strain', 'main_stage'], axis=1)
genes_avg_percentile = genes_orange_avg_scaled.quantile(q=percentile, axis=0)
genes_orange_avg_scaled = genes_orange_avg_scaled - genes_avg_percentile
genes_avg_percentile = genes_avg_percentile.replace(0, 1)
genes_orange_avg_scaled = genes_orange_avg_scaled / genes_avg_percentile
# Bound at max 0.1 (to remove outliers)
genes_orange_avg_scaled[genes_orange_avg_scaled > max_val] = max_val
genes_orange_avg_scaled[['Strain', 'main_stage']] = averaged[['Strain', 'main_stage']]
genes_orange_avg_scaled['Group'] = [GROUPS[strain] for strain in genes_orange_avg_scaled['Strain']]
genes_orange_avg_scaled.index = [strain + '_' + main
                                 for strain, main in genes_orange_avg_scaled[['Strain', 'main_stage']].values]
genes_orange_avg_scaled.to_csv(pathStages + 'genes_averaged_orange_mainStage_scale' + str(percentile)[2:] +
                               'percentileMax' + str(max_val) + '.tsv', sep='\t')
# *************
# *** Average data chubs
genes_mb= pd.read_csv('/home/karin/Documents/timeTrajectories/data/from_huston/normalised_mediaBuffer.tsv', sep='\t', index_col=0)
conditions_mb = pd.read_csv('/home/karin/Documents/timeTrajectories/data/from_huston/conditions_mediaBuffer.tsv', sep='\t', index_col=None)

# Average expression
genes_group = genes_mb[conditions_mb.Measurment].copy().T
genes_group = genes_group.join(pd.DataFrame(conditions_mb[['Group', 'Time']].values,
                                            index=conditions_mb['Measurment'], columns=['Group', 'Time']))
averaged = genes_group.groupby(['Group', 'Time']).mean().reset_index()
# Sort
averaged['Group'] = pd.Categorical(averaged['Group'], categories=['media','buff'], ordered=True)
averaged = averaged.sort_values(['Group', 'Time'])

averaged.to_csv('/home/karin/Documents/timeTrajectories/data/from_huston/averaged_mediaBuffer.tsv', sep='\t')

# Scale
percentile = 0.99
max_val = 0.1
genes_orange_avg_scaled = averaged.copy()
genes_orange_avg_scaled = genes_orange_avg_scaled.drop(['Group', 'Time'], axis=1)
genes_avg_percentile = genes_orange_avg_scaled.quantile(q=percentile, axis=0)
genes_orange_avg_scaled = genes_orange_avg_scaled - genes_avg_percentile
genes_avg_percentile = genes_avg_percentile.replace(0, 1)
genes_orange_avg_scaled = genes_orange_avg_scaled / genes_avg_percentile
# Bound at max 0.1 (to remove outliers)
genes_orange_avg_scaled[genes_orange_avg_scaled > max_val] = max_val
genes_orange_avg_scaled[['Group', 'Time']] = averaged[['Group', 'Time']]
genes_orange_avg_scaled.index = [group + '_' + str(time)
                                 for group, time in genes_orange_avg_scaled[['Group', 'Time']].values]
genes_orange_avg_scaled.to_csv('/home/karin/Documents/timeTrajectories/data/from_huston/averaged_scale' + str(percentile)[2:] +
                               'percentileMax' + str(max_val) + 'mediaBuffer.tsv', sep='\t')


# *****************
# ******* Merge results from DESeq2 between neighbouring stages
combined = []
phenotypes = [p for p in PHENOTYPES if p != 'yem']
folder = 'AX4_keepNA'
# files=[f for f in glob.glob(path_de_neighbouring+folder  + '/DE' + "*.tsv")]
# for f in files:
for idx in range(len(phenotypes) - 1):
    # f_split=f.split('/')[-1].split('_')
    # stage1=f_split[3]
    # stage2 = f_split[1]
    stage1 = phenotypes[idx]
    stage2 = phenotypes[idx + 1]
    f = path_de_neighbouring + folder + '/DE_' + stage2 + '_ref_' + stage1 + '_padj_lFC.tsv'
    data = pd.read_table(f, index_col=0)[['log2FoldChange', 'pvalue', 'padj']]
    # Remove genes for which pvalue was not calculated (e.g. outliers, ...)
    data = data[~data['pvalue'].isna()]
    data.columns = [stage1 + '_' + stage2 + '_' + col for col in data.columns]
    combined.append(data)
combined = pd.concat(combined, axis=1, sort=True)

# Adjust FDR across all comparisons, as they are all looked at together. Use only non-na pvalues
pvals = []
for col in combined.columns:
    if 'pvalue' in col:
        for row in combined.index:
            pval = combined.at[row, col]
            if not np.isnan(pval):
                pvals.append({'row': row, 'col': col, 'pval': pval})
pvals = pd.DataFrame(pvals)
pvals['FDR_overall'] = multipletests(pvals['pval'], method='fdr_bh')[1]
for row_name, data in pvals.iterrows():
    comparison = data['col'].rstrip('pvalue')
    combined.at[data['row'], comparison + 'FDR_overall'] = data['FDR_overall']

combined.to_csv(path_de_neighbouring + folder + '/combined.tsv', sep='\t')

# ***********
# *** Prepare stage annotation metafile for supplementary
stage_anno = conditions.copy()
stage_anno = stage_anno[[col for col in stage_anno.columns if col in PHENOTYPES or col == 'main_stage']]
stage_anno['Sample'] = [rep + '_hr' + str(t).zfill(2) for rep, t in zip(conditions['Replicate'], conditions['Time'])]
stage_anno['main_stage'] = stage_anno['main_stage'].fillna('NA')
for idx in stage_anno.index:
    if all(stage_anno.loc[idx, PHENOTYPES] == 0):
        stage_anno.loc[idx, PHENOTYPES] = 'NA'
stage_anno.to_csv('/home/karin/Documents/timeTrajectories/data/read_meta/stage_annotation.tsv', sep='\t', index=False)
